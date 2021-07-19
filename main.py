from openpyxl import load_workbook
from openpyxl import Workbook
from bs4 import BeautifulSoup
from fuzzywuzzy import fuzz
# from fuzzywuzzy import process
# import operator

# openpyxl - allows easy reading and writing to Excel spreadsheets
# BeautifulSoup - allows reading from xml files
# fuzzywuzzy - has functions that allow for fuzzy matching of words

# This function takes the old indicators spreadsheet and makes a new spreadsheet by
# organizing by statement type, and only showing indicators with a count higher
# than the number specified in the parameters


def new_indicators(statement_type, count, read_sheet_input, write_workbook_input):
    # First reads the sheet into a list
    modified_indicators = []

    for row in read_sheet_input.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):
        # If the count is lower than the one specified, we do not want to read in any more indicators
        if int(row[3]) < count:
            break
        if row[0] == statement_type and int(row[3]) > count:
            modified_indicators.append(row)

    # Then writes the list into a new Excel file
    write_sheet = write_workbook_input.create_sheet(statement_type)

    i = 1
    for row in modified_indicators:
        j = 1
        for cell in row:
            write_sheet.cell(row=i, column=j).value = cell
            j += 1
        i += 1

    new_filename = "new_indicators.xlsx"
    write_workbook_input.save(filename=new_filename)
    return modified_indicators

# Takes in the indicators of a specific financial statement and compares them with the elements
# of the taxonomy
# split_matches (True/False) - do we want the matches for each word split into individual cells or not
# num_matches - how many matches for each word do we want to display


def compare_taxonomy(statement_type, indicators, split_matches, num_matches):
    sheet = write_workbook[statement_type]

    # First iterates through the indicators list
    indicator_num = 1
    for row in indicators:
        category = row[1]
        indicator = row[2]
        category_indicator = category + " " + indicator

        element_num = 1
        ci_list = []
        c_list = []
        i_list = []

        # Then for each category/indicator, iterates through the elements of the taxonomy to find the best match
        for e in elements:
            # e[0] is the element/word
            element = e[0]
            # If we want to clean the word string before fuzzy matching is done, we can do so in this function
            new_element = clean_caption(element)
            # We want to split the string into its individual words based on where the capitals are
            # because I think the fuzzy matching works better with this
            new_element = split_capitals(element)

            # We make a list for each of these three words, with each row including the element word,
            # its number in the taxonomy, its fuzzy match score, and the difference in length of the element to the
            # corresponding category/indicator word
            ci_list.append([element, element_num, is_equal_num(category_indicator, new_element),
                            abs(len(element)-len(category_indicator))])
            c_list.append([element, element_num, is_equal_num(category, new_element), abs(len(element)-len(category))])
            i_list.append([element, element_num, is_equal_num(indicator, new_element),
                           abs(len(element)-len(indicator))])

            element_num += 1

        # We sort the list first by its fuzzy match score (highest first) and then by the difference in length of the
        # words (smallest first) because smaller difference in length implies the words are closer to being to the
        # same
        ci_list = sorted(ci_list, key=lambda x: (-x[2], x[3]))
        c_list = sorted(c_list, key=lambda x: (-x[2], x[3]))
        i_list = sorted(i_list, key=lambda x: (-x[2], x[3]))
        ci_list = ci_list[0:num_matches]
        c_list = c_list[0:num_matches]
        i_list = i_list[0:num_matches]

        if split_matches:
            for i in range(1, num_matches + 1):
                sheet.cell(row=indicator_num, column=4+i).value = str(ci_list[i-1][0]) + "," + str(ci_list[i-1][1]) + \
                                                                  "," + str(ci_list[i-1][2])
            for i in range(1, num_matches + 1):
                sheet.cell(row=indicator_num, column=4+num_matches+i).value = str(c_list[i-1][0]) + "," + \
                                                                        str(c_list[i-1][1]) + "," + str(c_list[i-1][2])
            for i in range(1, num_matches + 1):
                sheet.cell(row=indicator_num, column=4+(num_matches*2)+i).value = str(i_list[i-1][0]) + "," \
                                                                    + str(i_list[i-1][1]) + "," + str(i_list[i-1][2])

        else:
            new_ci_list = []
            for item in ci_list:
                new_ci_list.append("(" + item[0] + "," + str(item[1]) + "," + str(item[2]) + ")")
            new_c_list = []
            for item in c_list:
                new_c_list.append("(" + item[0] + "," + str(item[1]) + "," + str(item[2]) + ")")
            new_i_list = []
            for item in i_list:
                new_i_list.append("(" + item[0] + "," + str(item[1]) + "," + str(item[2]) + ")")
            sheet.cell(row=indicator_num, column=5).value = (",".join(new_ci_list))
            sheet.cell(row=indicator_num, column=6).value = (",".join(new_c_list))
            sheet.cell(row=indicator_num, column=7).value = (",".join(new_i_list))

        indicator_num += 1

    write_workbook.save(filename="new_indicators.xlsx")

# Similar kind of function as compare_taxonomy, except this time it compares the indicators
# of a specific financial statement to the words in the relevant part of the Uniform Chart of Accounts


def compare_accounts(statement_type, indicators, uniform, split_matches, num_matches):
    # Correct sheet in new_indicators spreadsheet to write to
    sheet = write_workbook[statement_type]

    # Goes through each indicator and sees if there is a match in Uniform Chart of Accounts list
    indicator_num = 1
    for row in indicators:
        category = row[1]
        indicator = row[2]
        category_indicator = category + " " + indicator

        ci_list = []
        c_list = []
        i_list = []

        for u in uniform:

            # There are now three lines of code for each category/indicator word because the matching word could be in
            # any of three columns of the UCA spreadsheet
            ci_list.append([u[2], u[4], is_equal_num(category_indicator, clean_caption(u[2])),
                            abs(len(u[2]) - len(category_indicator))])
            ci_list.append([u[3], u[4], is_equal_num(category_indicator, clean_caption(u[3])),
                            abs(len(u[3]) - len(category_indicator))])
            ci_list.append([u[1], u[4], is_equal_num(category_indicator, clean_caption(u[1])),
                            abs(len(u[1]) - len(category_indicator))])

            c_list.append([u[2], u[4], is_equal_num(category, clean_caption(u[2])), abs(len(u[2]) - len(category))])
            c_list.append([u[3], u[4], is_equal_num(category, clean_caption(u[3])), abs(len(u[3]) - len(category))])
            c_list.append([u[1], u[4], is_equal_num(category, clean_caption(u[1])), abs(len(u[1]) - len(category))])

            i_list.append([u[2], u[4], is_equal_num(indicator, clean_caption(u[2])), abs(len(u[2]) - len(indicator))])
            i_list.append([u[3], u[4], is_equal_num(indicator, clean_caption(u[3])), abs(len(u[3]) - len(indicator))])
            i_list.append([u[1], u[4], is_equal_num(indicator, clean_caption(u[1])), abs(len(u[1]) - len(indicator))])

        ci_list = sorted(ci_list, key=lambda x: (-x[2], x[3]))
        c_list = sorted(c_list, key=lambda x: (-x[2], x[3]))
        i_list = sorted(i_list, key=lambda x: (-x[2], x[3]))
        ci_list = ci_list[0:num_matches]
        c_list = c_list[0:num_matches]
        i_list = i_list[0:num_matches]

        if split_matches:
            for i in range(1, num_matches + 1):
                sheet.cell(row=indicator_num, column=4+(3*num_matches)+i).value = str(ci_list[i-1][0]) \
                                                            + "," + str(ci_list[i-1][1]) + "," + str(ci_list[i-1][2])
            for i in range(1, num_matches + 1):
                sheet.cell(row=indicator_num, column=4+(4*num_matches)+i).value = str(c_list[i-1][0]) \
                                                                + "," + str(c_list[i-1][1]) + "," + str(c_list[i-1][2])
            for i in range(1, num_matches + 1):
                sheet.cell(row=indicator_num, column=4+(5*num_matches)+i).value = str(i_list[i-1][0]) \
                                                                + "," + str(i_list[i-1][1]) + "," + str(i_list[i-1][2])

        else:
            new_ci_list = []
            for item in ci_list:
                new_ci_list.append("(" + item[0] + "," + str(item[1]) + "," + str(item[2]) + ")")

            new_c_list = []
            for item in c_list:
                new_c_list.append("(" + item[0] + "," + str(item[1]) + "," + str(item[2]) + ")")

            new_i_list = []
            for item in i_list:
                new_i_list.append("(" + item[0] + "," + str(item[1]) + "," + str(item[2]) + ")")
            sheet.cell(row=indicator_num, column=8).value = (",".join(new_ci_list))
            sheet.cell(row=indicator_num, column=9).value = (",".join(new_c_list))
            sheet.cell(row=indicator_num, column=10).value = (",".join(new_i_list))

        indicator_num += 1

    # Adds column titles to new sheet
    # Column titles depend on whether we want to display all three matches in
    # separate Excel cells (split_matches == True) or we want all three matches in
    # one Excel cell (split_matches == False

    sheet.insert_rows(idx=1)
    sheet.cell(row=1, column=1).value = "Statement Type"
    sheet.cell(row=1, column=2).value = "Category"
    sheet.cell(row=1, column=3).value = "Indicator"
    sheet.cell(row=1, column=4).value = "Count"

    if split_matches:
        for i in range(1, num_matches + 1):
            sheet.cell(row=1, column=4+i).value = "C-I Taxonomy Match " + str(i)
        for i in range(1, num_matches + 1):
            sheet.cell(row=1, column=4+num_matches+i).value = "C Taxonomy Match " + str(i)
        for i in range(1, num_matches + 1):
            sheet.cell(row=1, column=4+(2*num_matches)+i).value = "I Taxonomy Match " + str(i)
        for i in range(1, num_matches + 1):
            sheet.cell(row=1, column=4+(3*num_matches)+i).value = "C-I UCA Match " + str(i)
        for i in range(1, num_matches + 1):
            sheet.cell(row=1, column=4+(4*num_matches)+i).value = "C UCA Match " + str(i)
        for i in range(1, num_matches + 1):
            sheet.cell(row=1, column=4+(5*num_matches)+i).value = "I UCA Match " + str(i)

    else:
        sheet.cell(row=1, column=5).value = "C-I Taxonomy"
        sheet.cell(row=1, column=6).value = "C# Taxonomy"
        sheet.cell(row=1, column=7).value = "I# in Taxonomy"
        sheet.cell(row=1, column=8).value = "C-I# UCA"
        sheet.cell(row=1, column=9).value = "C# UCA"
        sheet.cell(row=1, column=10).value = "I# UCA"

    write_workbook.save(filename="new_indicators.xlsx")


# We want a function to look at each of the items in the UCA, and see if it has a matching item
# in the taxonomy. The purpose of this is to determine if the taxonomy is missing anything
def compare_taxonomy_accounts(sheet_name, elements_in):
    sheet = uniform_chart_workbook[sheet_name]

    if sheet_name == "Activities":
        col_length = 4
    else:
        col_length = 3

    row_num = 2
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=col_length, values_only=True):
        specific = row[1]
        small_category = row[2]
        if sheet_name == "Activities":
            large_category = row[3]
            lc_list = []

        s_list = []
        sc_list = []

        element_num = 1

        for item in elements_in:
            element = item[0]
            new_element = clean_caption(element)
            new_element = split_capitals(new_element)
            s_list.append([element, element_num, is_equal_num(specific, new_element), abs(len(element) - len(specific))])
            sc_list.append([element, element_num, is_equal_num(small_category, new_element), abs(len(element) - len(small_category))])
            if sheet_name == "Activities":
                lc_list.append([element, element_num, is_equal_num(large_category, new_element), abs(len(element) - len(large_category))])

            element_num += 1

        s_list = sorted(s_list, key=lambda x: (-x[2], x[3]))
        sc_list = sorted(sc_list, key=lambda x: (-x[2], x[3]))

        sheet.cell(row=row_num, column=col_length+1).value = (str(s_list[0][0]) + "," + str(s_list[0][1]) + "," + str(s_list[0][2]))
        sheet.cell(row=row_num, column=col_length + 2).value = (str(sc_list[0][0]) + "," + str(sc_list[0][1]) + "," + str(sc_list[0][2]))

        if sheet_name == "Activities":
            lc_list = sorted(lc_list, key=lambda x: (-x[2], x[3]))
            sheet.cell(row=row_num, column=col_length + 3).value = (str(lc_list[0][0]) + "," + str(lc_list[0][1]) + "," + str(lc_list[0][2]))

        row_num += 1

    uniform_chart_workbook.save(filename="uniform_chart.xlsx")



# Sometimes doing string cleaning/preprocessing before the fuzzy matching takes place can yield better results.
# I determined it doesn't, but some ways in which it theoretically might have are commented out in the function


def clean_caption(old_word):
    # old_word = old_word.lower()
    # words = old_word.split()
    # Removes trailing 's' because sometimes that is a source of mismatch
    # i = 0
    # for w in words:
    #    words[i] = w.rstrip('s')
    #    i += 1
    # new_word = "".join(words)
    # new_word = new_word.replace('-', '')
    # new_word = new_word.replace(':', '')
    # new_word = new_word.replace('member', '')
    # new_word = new_word.strip()
    # new_word = new_word.strip('\n')
    # return new_word
    return old_word

# Splits string into individual words that start with the capital letters found in the string


def split_capitals(word):
    new_word = ""
    for letter in word:
        if letter.isupper():
            new_word += (" " + letter)
        else:
            new_word += letter
    new_word.strip()
    return new_word

# Returns true/false based on if the fuzzy match score meets a certain threshold. I decided not to use this function


def is_equal(word1, word2):
    # word1 = clean_caption(word1)
    # word2 = clean_caption(word2)
    # ratio = fuzz.ratio(word1, word2)
    # partial_ratio = fuzz.partial_ratio(word1, word2)
    # token_sort_ratio = fuzz.token_sort_ratio(word1, word2)
    token_set_ratio = fuzz.token_set_ratio(word1, word2)

    if token_set_ratio > 70:
        return True
    return False

# Returns the fuzzy match score of two words. As you can see above and below, fuzzywuzzy offers many different
# ways to do funny match, I determined token_set_ratio was the most effective for our purposes


def is_equal_num(word1, word2):
    token_set_ratio = fuzz.token_set_ratio(word1, word2)
    # partial_ratio = fuzz.partial_ratio(word1, word2)
    # ratio = fuzz.ratio(word1, word2)

    return token_set_ratio

# Makes it easier to input information into an Excel cell, I had it used this function for an earlier version of the
# code but don't need it in this version


def cell_input(column_num, sheet_in, indicator_num_in, element_num_in):
    cell_in = sheet_in.cell(row=indicator_num_in, column=column_num).value
    if cell_in is not None:
        sheet_in.cell(row=indicator_num_in, column=column_num).value = (cell_in + str(element_num_in) + ",")
    else:
        sheet_in.cell(row=indicator_num_in, column=column_num).value = (str(element_num_in) + ",")


if __name__ == '__main__':
    '''
    # Creates a workbook to read from the indicators Excel file
    read_workbook = load_workbook(filename="indicators.xlsx")
    read_sheet = read_workbook["Statement Agg"]

    # Creates a workbook to write to an Excel file
    write_workbook = Workbook()

    # Writes to a new Excel file with only relevant indicators, one sheet per financial statement
    activities_indicators = new_indicators("Statement of Activities", 10, read_sheet, write_workbook)
    balance_indicators = new_indicators("Balance Sheet", 10, read_sheet, write_workbook)
    revenues_indicators = new_indicators("Statement of Revenues Expenditures and Changes in Fund Balance", 10,
                                         read_sheet, write_workbook)
    governmental_funds_indicators = new_indicators("Statement of Net Position Governmental Funds", 10, read_sheet,
                                                   write_workbook)
    proprietary_funds_indicators = new_indicators("Statement of Net Position Proprietary Funds", 10, read_sheet,
                                                  write_workbook)
    rec_indicators = new_indicators("Statement of Revenues Expenses and Changes in Net Position", 10, read_sheet,
                                    write_workbook)
    cash_flows_indicators = new_indicators("Statement of Cash Flows", 10, read_sheet,
                                    write_workbook)
    '''

    # Create a Beautiful Soup object from the xml/xsd file
    with open("taxonomy.xsd") as fp:
        soup = BeautifulSoup(fp, "xml")

    # Finds all definitions in the taxonomy and adds them to a list
    # Note - not sure tactically what the difference is between the "definitions" and the "elements"
    # and other parts of the taxonomy, a question for Marc
    raw_definitions = soup.find_all('definition')
    definitions = []
    for d in raw_definitions:
        definitions.append(d.string)

    # Parses the previous definitions list and makes a new cleaner version
    new_definitions = []
    for d in definitions:
        j = d.split(" - ")
        k = []
        for i in j:
            i = i.strip()
            if i != ",":
                k.append(i)
        new_definitions.append(k)

    # Writes definitions list into an Excel file
    taxonomy_workbook = Workbook()
    sheet = taxonomy_workbook.active
    i = 1
    for row in new_definitions:
        j = 1
        for cell in row:
            sheet.cell(row=i, column=j).value = cell
            j += 1
        i += 1

    # Finds all elements in the taxonomy and adds them to a list
    elements = []
    raw_elements = soup.find_all('element')
    for e in raw_elements:
        f = []
        f.append(e['name'])
        f.append(e['abstract'])
        elements.append(f)

    # Writes elements list into an Excel file
    elements_sheet = taxonomy_workbook.create_sheet("elements")
    i = 1
    for row in elements:
        j = 1
        for cell in row:
            elements_sheet.cell(row=i, column=j).value = cell
            j += 1
        i += 1

    new_filename = "taxonomy_indicators.xlsx"
    taxonomy_workbook.save(filename=new_filename)

    '''

    # Figures out how indicators list compares to taxonomy list
    compare_taxonomy("Statement of Activities", activities_indicators, True, 4)
    compare_taxonomy("Balance Sheet", balance_indicators, True, 4)
    compare_taxonomy("Statement of Revenues Expenditures and Changes in Fund Balance", revenues_indicators, True, 4)
    compare_taxonomy("Statement of Revenues Expenses and Changes in Net Position", rec_indicators, True, 4)
    compare_taxonomy("Statement of Net Position Governmental Funds", governmental_funds_indicators, True, 4)
    compare_taxonomy("Statement of Net Position Proprietary Funds", proprietary_funds_indicators, True, 4)
    compare_taxonomy("Statement of Cash Flows", cash_flows_indicators, True, 4)
    '''


    # Creates a workbook to read from the Uniform Chart of Accounts Excel file
    uniform_chart_workbook = load_workbook(filename="uniform_chart.xlsx")
    uniform_activities = uniform_chart_workbook["Activities"]
    uniform_bre = uniform_chart_workbook["Combined B,R,E"]

    '''
    # Takes the Uniform Chart of Accounts Excel Spreadsheet and makes it into list with cleaned up values
    # This is one for the Activities sheet in this file
    uniform_activities_list = []
    for row in uniform_activities.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):
        new_row = []
        for cell in row:
            new_row.append(cell)
        new_row.append("A" + str(new_row[0]))
        uniform_activities_list.append(new_row)

    # IF WE WANT TO USE THE REVISED UCA EXCEL, WE CAN USE THE INDIVIDUAL B,R, AND E
    # SHEETS INSTEAD OF THIS COMBINED ONE
    # This one is for the "Combined B,R,E" sheet in this file
    uniform_bre_list = []
    for row in uniform_bre.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
        new_row = []
        for cell in row:
            new_row.append(cell)
        new_row.append(" ")
        new_row.append("BRE" + str(new_row[0]))
        uniform_bre_list.append(new_row)

    # This one is the previous two combined
    combined_list = []
    for row in uniform_activities_list:
        combined_list.append(row)
    for row in uniform_bre_list:
        combined_list.append(row)

    # Some statement types only take in the activities_list, some only take the bre_list, and some take the
    # combined list, this could be changed depending on what is deemed relevant
    # Figures out how indicators list compares to Uniform Chart of Accounts
    compare_accounts("Statement of Activities", activities_indicators, uniform_activities_list, True, 4)
    compare_accounts("Balance Sheet", balance_indicators, uniform_bre_list, True, 4)
    compare_accounts("Statement of Revenues Expenditures and Changes in Fund Balance", revenues_indicators,
                     uniform_bre_list, True, 4)
    compare_accounts("Statement of Revenues Expenses and Changes in Net Position", rec_indicators,
                     combined_list, True, 4)
    compare_accounts("Statement of Net Position Governmental Funds", governmental_funds_indicators,
                     combined_list, True, 4)
    compare_accounts("Statement of Net Position Proprietary Funds", proprietary_funds_indicators,
                     combined_list, True, 4)
    compare_accounts("Statement of Cash Flows", cash_flows_indicators, combined_list, True, 4)
    '''

    compare_taxonomy_accounts("Activities", elements)
    compare_taxonomy_accounts("Combined B,R,E", elements)
    compare_taxonomy_accounts("Funds", elements)